from django.shortcuts import render, get_object_or_404
from django.http import FileResponse, HttpResponse
from django.conf import settings
from django.core.files.base import ContentFile
import os
import time
from io import BytesIO
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment
from .utils import get_group_type_excel_data, get_individual_type_excel_data, get_encoded_reg_id, get_decoded_reg_id
from fest.pdf_generators.entrypass.entrypass_generator import render_entrypass
from fest.models import Registration
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated


def admin_required(view_func):
    def inner(request, *args, **kwargs):
        if request.user.is_authenticated:
            return view_func(request, *args, **kwargs)
        else:
            return HttpResponse("Access Denied")
    return inner
        
def verify_registration(request, reg_code):
    pk = None
    try:
        pk = get_decoded_reg_id(reg_code)
    except Exception as e:
        pass
    if pk:
        reg = Registration.objects.filter(pk=pk, is_approved=True).first()
    else:
        reg = None
    return render(request, 'fest/verify_reg.html', context={'reg': reg})
     

def download_asset(request, filename):
    filepath = settings.BASE_DIR / ('frontend/main_site/static/assets/' + filename)
    print(filepath)
    if os.path.exists(filepath):
        return FileResponse(open(filepath, 'rb'), filename=filename)
    return HttpResponse(f"File not found!")


@api_view()
@permission_classes([IsAuthenticated])
def download_response_excel(request):
    contest = request.GET.get('contest', 'all')
    approval = request.GET.get('approval', 'all')
    if contest in ['all', 'poster', 'lfr']:
        data = get_group_type_excel_data(contest, approval)
    else:
        data = get_individual_type_excel_data(contest, approval)
    workbook = Workbook()
    worksheet = workbook.active
    for row_index, row_data in enumerate(data):
        for column_index, cell_value in enumerate(row_data):
            worksheet.cell(row=row_index + 1, column=column_index + 1, value=cell_value)
            
    # Stylings
    num_cols = worksheet.max_column
    for i in range(num_cols):
        worksheet.column_dimensions[chr(ord('A')+i)].width = 20
        if i > 25:
            break
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    buffer = BytesIO()
    workbook.save(buffer)
    filename = f'Response {contest.upper()} {datetime.now().strftime("%H:%M:%S %Y-%m-%d")}.xlsx'
    return FileResponse(
        ContentFile(buffer.getvalue()),
        content_type='application/vnd.ms-excel', 
        filename=filename, as_attachment=True
    )
    
    
def download_entrypass(request, reg_code):
    try:
        pk = get_decoded_reg_id(reg_code)
    except Exception as e:
        return HttpResponse("Registration Unknown")
    reg = get_object_or_404(Registration, pk=pk)
    if not reg.is_approved:
        return HttpResponse("Registration is not approved")
    doc_pdf = render_entrypass(request, reg)
    filename = f"Technoventure3.0 Entrypass.pdf"
    return FileResponse(ContentFile(doc_pdf), filename=filename)
    